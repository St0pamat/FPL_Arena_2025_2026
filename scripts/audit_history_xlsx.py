"""Audit history25-26 xlsx vs existing player data."""
import glob
import json
import os
import re
from collections import defaultdict

import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HIGHLIGHTS_DIR = os.path.join(BASE, "highlights")
PLAYERS_TS = os.path.join(BASE, "src", "data", "players.ts")
HIGHLIGHTS_JSON = os.path.join(BASE, "player_highlights.json")


def load_players_ts():
    text = open(PLAYERS_TS, encoding="utf-8").read()
    players = {}
    for m in re.finditer(
        r"id:\s*(\d+).*?manager:\s*\"([^\"]+)\".*?team:\s*\"([^\"]+)\".*?avgPosition:\s*([\d.]+).*?"
        r"winStreak:\s*\"([^\"]*)\".*?monthlyWins:\s*\"([^\"]*)\".*?weeksTop:\s*(\d+).*?weeksBottom:\s*(\d+)",
        text,
        re.DOTALL,
    ):
        pid = int(m.group(1))
        players[pid] = {
            "manager": m.group(2),
            "team": m.group(3),
            "avgPosition": float(m.group(4)),
            "winStreak": m.group(5),
            "monthlyWins": m.group(6),
            "weeksTop": int(m.group(7)),
            "weeksBottom": int(m.group(8)),
        }
    return players


def team_info(wb):
    info = {}
    if "Team Info" not in wb.sheetnames:
        return info
    for row in wb["Team Info"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            info[row[0]] = row[1]
    return info


def sheet_headers(wb, name):
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    return [c.value for c in ws[1]]


def main():
    history_files = sorted(
        f
        for f in glob.glob(os.path.join(HIGHLIGHTS_DIR, "*.xlsx"))
        if "history" in os.path.basename(f).lower()
    )
    print(f"History files: {len(history_files)}")

    all_sheets = defaultdict(set)
    team_keys = defaultdict(set)
    per_file = {}

    for path in history_files:
        fname = os.path.basename(path)
        pid = re.match(r"(\d+)", fname).group(1)
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheets = wb.sheetnames
        for s in sheets:
            all_sheets[s].add(pid)
        info = team_info(wb)
        for k in info:
            team_keys[k].add(str(info[k])[:80])
        headers_by_sheet = {s: sheet_headers(wb, s) for s in sheets}
        per_file[int(pid)] = {
            "file": fname,
            "sheets": sheets,
            "team_info": info,
            "headers": headers_by_sheet,
        }
        wb.close()

    print("\n=== SHEETS ===")
    for s in sorted(all_sheets):
        print(f"  {s} ({len(all_sheets[s])} files)")

    print("\n=== TEAM INFO KEYS ===")
    for k in sorted(team_keys):
        sample = next(iter(team_keys[k]))
        print(f"  {k}: e.g. {sample}")

    # Compare history-only sheets vs konwertuj_highlights sheets
    standard = {
        "Team Info",
        "GW Info",
        "GW Picks",
        "Player Gain - Loss",
        "Transfer Quality",
        "Gain Analysis",
        "Loss Analysis",
        "Formations",
        "Event Points Total",
        "Exp vs Real Performance",
        "Target Event Ratios",
        "Team of the Year",
    }
    history_only = set(all_sheets) - standard
    print("\n=== SHEETS NOT IN konwertuj_highlights.py ===")
    for s in sorted(history_only):
        p = next(iter(per_file.values()))
        if s in p["headers"]:
            print(f"  {s}: {p['headers'][s]}")

    # Sample history-only sheet content for 22952
    pid = 22952
    if pid in per_file:
        d = per_file[pid]
        print(f"\n=== SAMPLE CONTENT pid={pid} ===")
        for s in sorted(history_only):
            path = next(f for f in history_files if str(pid) in os.path.basename(f))
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            if s not in wb.sheetnames:
                wb.close()
                continue
            ws = wb[s]
            rows = list(ws.iter_rows(max_row=8, values_only=True))
            print(f"\n--- {s} ---")
            print("  headers:", rows[0] if rows else [])
            for r in rows[1:6]:
                print(" ", r)
            wb.close()

    # Load highlights json for comparison
    with open(HIGHLIGHTS_JSON, encoding="utf-8") as f:
        hl = json.load(f)

    players = load_players_ts()
    print("\n=== FPL SEASON OR vs bestOr ===")
    for pid, p in sorted(players.items()):
        h = hl.get(str(pid), {})
        season_or = h.get("seasonOr")
        best_or = p.get("bestOr") if "bestOr" in p else None
        # parse bestOr from players.ts manually for this pid
        text = open(PLAYERS_TS, encoding="utf-8").read()
        block = re.search(rf"id:\s*{pid},.*?bestOr:\s*\"([^\"]+)\"", text, re.DOTALL)
        best_or = block.group(1).replace(" ", "") if block else "?"
        or_match = season_or and str(season_or) == best_or.replace(" ", "")
        print(f"  {pid} {p['manager'][:20]:20} seasonOr={season_or} bestOr={best_or} match={or_match}")

    print("\n=== GW POINTS TOTAL vs H2H score ===")
    for pid, p in sorted(players.items()):
        h = hl.get(str(pid), {})
        gw_sum = sum(g["points"] for g in h.get("gwPoints", []))
        text = open(PLAYERS_TS, encoding="utf-8").read()
        block = re.search(rf"id:\s*{pid},.*?score:\s*(\d+)", text, re.DOTALL)
        h2h_score = int(block.group(1)) if block else 0
        diff = gw_sum - h2h_score
        if abs(diff) > 0:
            print(f"  {pid}: FPL sum={gw_sum} H2H score={h2h_score} diff={diff}")

    # Extract unique metrics from history sheets
    print("\n=== UNIQUE METRICS IN HISTORY FILES ===")
    metrics = {}
    for path in history_files:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        info = team_info(wb)
        pid = int(re.match(r"(\d+)", os.path.basename(path)).group(1))
        metrics[pid] = info
        wb.close()

    all_metric_keys = sorted({k for m in metrics.values() for k in m})
    print("Team Info fields:", all_metric_keys)


if __name__ == "__main__":
    main()
