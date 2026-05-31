"""Extract history25-26 xlsx + compute league stats for comparison."""
import glob
import json
import os
import re
from collections import defaultdict

import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HIGHLIGHTS_DIR = os.path.join(BASE, "highlights")
WYNIKI = os.path.join(BASE, "wyniki_meczy.json")
PLAYERS_TS = os.path.join(BASE, "src", "data", "players.ts")
OUT = os.path.join(BASE, "scripts", "audit_output.json")


def parse_players():
    text = open(PLAYERS_TS, encoding="utf-8").read()
    players = {}
    for m in re.finditer(
        r"id:\s*(\d+).*?manager:\s*\"([^\"]+)\".*?team:\s*\"([^\"]+)\".*?"
        r"score:\s*(\d+).*?avgPosition:\s*([\d.]+).*?"
        r"superStar:\s*\"([^\"]+)\".*?rankKiller:\s*\"([^\"]+)\".*?"
        r"winStreak:\s*\"([^\"]*)\".*?monthlyWins:\s*\"([^\"]*)\".*?"
        r"weeksTop:\s*(\d+).*?weeksBottom:\s*(\d+)",
        text,
        re.DOTALL,
    ):
        pid = int(m.group(1))
        players[pid] = {
            "manager": m.group(2),
            "team": m.group(3),
            "score": int(m.group(4)),
            "avgPosition": float(m.group(5)),
            "superStar": m.group(6),
            "rankKiller": m.group(7),
            "winStreak": m.group(8),
            "monthlyWins": m.group(9),
            "weeksTop": int(m.group(10)),
            "weeksBottom": int(m.group(11)),
        }
    return players


def find_gw_history_sheet(wb, pid):
    target = f"{pid} GW History"
    if target in wb.sheetnames:
        return target
    for s in wb.sheetnames:
        if "GW History" in s:
            return s
    return None


def parse_history(path):
    fname = os.path.basename(path)
    pid = int(re.match(r"(\d+)", fname).group(1))
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = find_gw_history_sheet(wb, pid)
    if not sheet:
        wb.close()
        return pid, None

    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    summary = {}
    if rows:
        hdr = rows[0]
        val = rows[1] if len(rows) > 1 else ()
        for i, k in enumerate(hdr):
            if k and i < len(val):
                summary[k] = val[i]

    gw_hdr_idx = None
    for i, row in enumerate(rows):
        if row and row[0] == "GW":
            gw_hdr_idx = i
            break

    gw_rows = []
    if gw_hdr_idx is not None:
        headers = rows[gw_hdr_idx]
        for row in rows[gw_hdr_idx + 1 :]:
            if not row or row[0] is None:
                continue
            try:
                int(row[0])
            except (TypeError, ValueError):
                continue
            gw_rows.append(dict(zip(headers, row)))

    # Player Analysis
    wb2 = openpyxl.load_workbook(path, data_only=True, read_only=True)
    pa = []
    if "Player Analysis" in wb2.sheetnames:
        pa_rows = list(wb2["Player Analysis"].iter_rows(values_only=True))
        hdr_row = None
        for i, row in enumerate(pa_rows):
            if row and row[0] == "Player":
                hdr_row = i
                break
        if hdr_row is not None:
            for row in pa_rows[hdr_row + 1 : hdr_row + 6]:
                if row and row[0]:
                    pa.append({
                        "helper": row[0],
                        "helperGain": row[1],
                        "hurter": row[3],
                        "hurterLoss": row[4],
                        "mostPts": row[6],
                        "mostPtsVal": row[7],
                        "mostCap": row[9],
                        "mostCapVal": row[10],
                        "mostStart": row[12],
                        "mostStartVal": row[13],
                        "mostBench": row[15],
                        "mostBenchVal": row[16],
                    })
    wb2.close()

    # derived stats
    vs_top10k = []
    vs_overall = []
    best_gw_rank = None
    worst_gw_rank = None
    green_arrows = 0
    red_arrows = 0
    for g in gw_rows:
        pts = float(g.get("Points") or 0)
        t10 = float(g.get("Top10k Avg") or 0)
        oa = float(g.get("Overall Avg") or 0)
        vs_top10k.append(pts - t10)
        vs_overall.append(pts - oa)
        rc = g.get("Rank Change %")
        if rc is not None and rc != 0:
            try:
                if float(rc) > 0:
                    green_arrows += 1
                elif float(rc) < 0:
                    red_arrows += 1
            except (TypeError, ValueError):
                pass
        gr = g.get("GW Rank")
        if gr:
            try:
                gr = int(gr)
                if best_gw_rank is None or gr < best_gw_rank:
                    best_gw_rank = gr
                if worst_gw_rank is None or gr > worst_gw_rank:
                    worst_gw_rank = gr
            except (TypeError, ValueError):
                pass

    return pid, {
        "file": fname,
        "summary": summary,
        "gwCount": len(gw_rows),
        "totalPoints": sum(float(g.get("Points") or 0) for g in gw_rows),
        "avgVsTop10k": round(sum(vs_top10k) / len(vs_top10k), 2) if vs_top10k else None,
        "avgVsOverall": round(sum(vs_overall) / len(vs_overall), 2) if vs_overall else None,
        "weeksAboveTop10k": sum(1 for v in vs_top10k if v > 0),
        "weeksBelowTop10k": sum(1 for v in vs_top10k if v < 0),
        "bestGwRank": best_gw_rank,
        "worstGwRank": worst_gw_rank,
        "greenArrowsFromHistory": green_arrows,
        "redArrowsFromHistory": red_arrows,
        "playerAnalysisTop": pa,
        "mostStarted": summary.get("Most started"),
        "mostBenched": summary.get("Most Benched"),
    }


def league_stats_from_matches():
    with open(WYNIKI, encoding="utf-8") as f:
        blocks = json.load(f)

    team_points_by_gw = defaultdict(dict)
    all_teams = set()
    for block in blocks:
        gw = block["gw"]
        for m in block.get("matches", []):
            ta, tb = m["teamA"], m["teamB"]
            pa, pb = int(m["pointsA"]), int(m["pointsB"])
            team_points_by_gw[gw][ta] = pa
            team_points_by_gw[gw][tb] = pb
            all_teams.add(ta)
            all_teams.add(tb)

    # weekly top scorer
    weekly_top = {}
    top_scorer_counts = defaultdict(int)
    for gw, teams in team_points_by_gw.items():
        if not teams:
            continue
        best_team = max(teams, key=teams.get)
        best_pts = teams[best_team]
        weekly_top[gw] = {"team": best_team, "points": best_pts}
        top_scorer_counts[best_team] += 1

    # standings history for weeks top/bottom and avg position
    from datetime import datetime

    # simple cumulative points table per gw
    teams = sorted(all_teams)
    cum = {t: 0 for t in teams}
    positions_by_team = defaultdict(list)
    for gw in sorted(team_points_by_gw.keys()):
        # update cumulative H2H league points from matches - need wyniki structure
        pass

    # Use buildStandingsHistory logic - compute from wyniki
    cum_pts = {t: {"pts": 0, "score": 0} for t in teams}
    for block in sorted(blocks, key=lambda b: b["gw"]):
        gw = block["gw"]
        for m in block.get("matches", []):
            ta, tb = m["teamA"], m["teamB"]
            pa, pb = int(m["pointsA"]), int(m["pointsB"])
            cum_pts[ta]["score"] += pa
            cum_pts[tb]["score"] += pb
            if pa > pb:
                cum_pts[ta]["pts"] += 3
            elif pb > pa:
                cum_pts[tb]["pts"] += 3
            else:
                cum_pts[ta]["pts"] += 1
                cum_pts[tb]["pts"] += 1
        ranked = sorted(cum_pts.items(), key=lambda x: (-x[1]["pts"], -x[1]["score"]))
        for pos, (team, _) in enumerate(ranked, 1):
            positions_by_team[team].append(pos)

    weeks_top = {t: sum(1 for p in positions_by_team[t] if p == 1) for t in teams}
    weeks_bottom = {t: sum(1 for p in positions_by_team[t] if p == len(teams)) for t in teams}
    avg_pos = {t: round(sum(positions_by_team[t]) / len(positions_by_team[t]), 2) for t in teams}

    # win/loss streaks from matches
    team_outcomes = defaultdict(list)
    for block in sorted(blocks, key=lambda b: b["gw"]):
        gw = block["gw"]
        for m in block.get("matches", []):
            ta, tb = m["teamA"], m["teamB"]
            pa, pb = int(m["pointsA"]), int(m["pointsB"])
            oa = "W" if pa > pb else ("L" if pa < pb else "D")
            ob = "W" if pb > pa else ("L" if pb < pa else "D")
            team_outcomes[ta].append((gw, oa))
            team_outcomes[tb].append((gw, ob))

    def longest(outcomes, pred):
        best = (0, None, None)
        cur = 0
        start = None
        for gw, o in outcomes:
            if pred(o):
                if cur == 0:
                    start = gw
                cur += 1
                if cur > best[0]:
                    best = (cur, start, gw)
            else:
                cur = 0
                start = None
        return best

    streaks = {}
    for t, oc in team_outcomes.items():
        wl = longest(oc, lambda o: o == "W")
        ll = longest(oc, lambda o: o == "L")
        streaks[t] = {"winStreak": wl, "lossStreak": ll}

    return {
        "weeklyTopScorerCounts": dict(top_scorer_counts),
        "weeksTopComputed": weeks_top,
        "weeksBottomComputed": weeks_bottom,
        "avgPositionComputed": avg_pos,
        "winLossStreaks": streaks,
        "topWeeklyScorers": sorted(top_scorer_counts.items(), key=lambda x: -x[1])[:5],
    }


def main():
    players = parse_players()
    team_by_id = {pid: p["team"] for pid, p in players.items()}
    id_by_team = {p["team"]: pid for pid, p in players.items()}

    history = {}
    for path in sorted(glob.glob(os.path.join(HIGHLIGHTS_DIR, "*history*.xlsx"))):
        pid, data = parse_history(path)
        if data:
            history[pid] = data

    league = league_stats_from_matches()

    discrepancies = []
    for pid, p in players.items():
        team = p["team"]
        h = history.get(pid, {})
        s = h.get("summary", {})

        # score vs history total
        hist_pts = h.get("totalPoints")
        if hist_pts and abs(hist_pts - p["score"]) > 5:
            discrepancies.append({
                "id": pid, "field": "score/totalPoints",
                "players.ts": p["score"], "history": hist_pts,
                "note": "H2H score vs FPL season sum (expected to differ slightly)",
            })

        # avg position
        comp = league["avgPositionComputed"].get(team)
        if comp and abs(comp - p["avgPosition"]) > 0.05:
            discrepancies.append({
                "id": pid, "field": "avgPosition",
                "players.ts": p["avgPosition"], "computed": comp,
            })

        # weeks top/bottom
        wt = league["weeksTopComputed"].get(team)
        wb = league["weeksBottomComputed"].get(team)
        if wt is not None and wt != p["weeksTop"]:
            discrepancies.append({"id": pid, "field": "weeksTop", "players.ts": p["weeksTop"], "computed": wt})
        if wb is not None and wb != p["weeksBottom"]:
            discrepancies.append({"id": pid, "field": "weeksBottom", "players.ts": p["weeksBottom"], "computed": wb})

        # win streak
        st = league["winLossStreaks"].get(team, {}).get("winStreak")
        if st and st[0]:
            ws = p["winStreak"]
            m = re.search(r"(\d+)", ws)
            stored = int(m.group(1)) if m else 0
            if stored != st[0]:
                discrepancies.append({
                    "id": pid, "field": "winStreak",
                    "players.ts": ws, "computed": f"{st[0]} (GW{st[1]}-GW{st[2]})",
                })

        # superStar vs history
        ss_hist = s.get("Super Star")
        if ss_hist and p["superStar"]:
            discrepancies.append({
                "id": pid, "field": "superStar",
                "players.ts": p["superStar"], "history": str(ss_hist)[:60],
                "note": "Different source/metric possible",
            })

    new_metrics = {
        "fromHistoryNotInUI": [
            "avgVsTop10k (season average vs elite managers)",
            "weeksAboveTop10k / weeksBelowTop10k",
            "bestGwRank / worstGwRank (best single-GW global rank)",
            "mostStarted player",
            "mostBenched player (count, not points)",
            "per-GW Captain, Main Gain, Main Damage",
            "Team Value progression",
            "Player Analysis ranked lists (top 5 helpers/hurters)",
        ],
        "fromScreenshotNotInUI": [
            "weeklyTopScorerCounts (how many GWs as league top scorer)",
            "weeksTop / weeksBottom leaderboards",
            "avgPosition ranking board",
            "Manager of Month detailed W/D/L (only month names in profiles)",
        ],
    }

    out = {
        "historyCount": len(history),
        "sampleHistory22952": history.get(22952),
        "leagueStats": league,
        "discrepancies": discrepancies,
        "newMetrics": new_metrics,
        "historySummaryFields": list(next(iter(history.values()))["summary"].keys()) if history else [],
    }

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(json.dumps(out, ensure_ascii=False, indent=2)[:8000])


if __name__ == "__main__":
    main()
