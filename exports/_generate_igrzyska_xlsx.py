# -*- coding: utf-8 -*-
"""One-off export: Igrzyska Kapci Kłapcia 2025/26 → Excel. Does not modify app code."""
from __future__ import annotations

import json
import re
from datetime import datetime
from functools import cmp_to_key
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = Path(__file__).resolve().parent / "Igrzyska_Kapci_Klapcia_Sezon_2025-26.xlsx"

matches_by_gw = json.loads((ROOT / "wyniki_meczy.json").read_text(encoding="utf-8"))
players_src = (ROOT / "src" / "data" / "players.ts").read_text(encoding="utf-8")

player_blocks = re.findall(r"\{\s*id:\s*(\d+),(.*?)(?=\n\s*\{|\n\];)", players_src, re.S)
players: dict[str, dict] = {}
for pid, body in player_blocks:

    def grab(key: str, cast=str):
        m = re.search(
            rf'(?<![A-Za-z_]){re.escape(key)}:\s*(?:"([^"]*)"|(-?\d+(?:\.\d+)?))',
            body,
        )
        if not m:
            return None
        raw = m.group(1) if m.group(1) is not None else m.group(2)
        return cast(raw)

    team = grab("team")
    if not team:
        continue
    players[team] = {
        "id": int(pid),
        "manager": grab("manager") or "",
        "discord": grab("discord") or "",
        "team": team,
        "rank": int(grab("rank", float) or 0),
        "gw19Rank": int(grab("gw19Rank", float) or 0),
        "w": int(grab("w", float) or 0),
        "d": int(grab("d", float) or 0),
        "l": int(grab("l", float) or 0),
        "score": int(grab("score", float) or 0),
        "pts": int(grab("pts", float) or 0),
        "seasons": int(grab("seasons", float) or 0),
        "bestOr": grab("bestOr") or "",
        "bestOrSeason": grab("bestOrSeason") or "",
        "avgPosition": float(grab("avgPosition", float) or 0),
        "transfers": int(grab("transfers", float) or 0),
        "hits": int(grab("hits", float) or 0),
        "greenArrows": int(grab("greenArrows", float) or 0),
        "captainPts": int(grab("captainPts", float) or 0),
        "mostCaptained": grab("mostCaptained") or "",
        "mostPointsPlayer": grab("mostPointsPlayer") or "",
        "superStar": grab("superStar") or "",
        "rankKiller": grab("rankKiller") or "",
        "pointsBenched": int(grab("pointsBenched", float) or 0),
        "winStreak": grab("winStreak") or "",
        "monthlyWins": grab("monthlyWins") or "",
        "bestGw": grab("bestGw") or "",
        "weeksTop": int(grab("weeksTop", float) or 0),
        "weeksBottom": int(grab("weeksBottom", float) or 0),
    }


def empty_row(team: str) -> dict:
    return {"team": team, "w": 0, "d": 0, "l": 0, "pts": 0, "score": 0}


def mini_league(teams, match_log):
    s = {t: {"pts": 0, "score": 0} for t in teams}
    for m in match_log:
        if m["teamA"] not in teams or m["teamB"] not in teams:
            continue
        s[m["teamA"]]["pts"] += m["leaguePtsA"]
        s[m["teamB"]]["pts"] += m["leaguePtsB"]
        s[m["teamA"]]["score"] += m["fa"]
        s[m["teamB"]]["score"] += m["fb"]
    return s


def compare(a, b, match_log):
    if b["pts"] != a["pts"]:
        return b["pts"] - a["pts"]
    if b["score"] != a["score"]:
        return b["score"] - a["score"]
    mini = mini_league([a["team"], b["team"]], match_log)
    if mini[b["team"]]["pts"] != mini[a["team"]]["pts"]:
        return mini[b["team"]]["pts"] - mini[a["team"]]["pts"]
    if mini[b["team"]]["score"] != mini[a["team"]]["score"]:
        return mini[b["team"]]["score"] - mini[a["team"]]["score"]
    if a["team"] < b["team"]:
        return -1
    if a["team"] > b["team"]:
        return 1
    return 0


def compute_standings(through_gw: int):
    team_names = list(players.keys())
    if not team_names:
        s = set()
        for block in matches_by_gw:
            for m in block["matches"]:
                s.add(m["teamA"])
                s.add(m["teamB"])
        team_names = sorted(s)

    stats = {t: empty_row(t) for t in team_names}
    match_log = []
    for block in sorted(matches_by_gw, key=lambda x: int(x["gw"])):
        gw = int(block["gw"])
        if gw > through_gw:
            break
        for m in block.get("matches") or []:
            for t in (m["teamA"], m["teamB"]):
                if t not in stats:
                    stats[t] = empty_row(t)
            fa, fb = int(m["pointsA"]), int(m["pointsB"])
            if fa > fb:
                lpa, lpb = 3, 0
                stats[m["teamA"]]["w"] += 1
                stats[m["teamB"]]["l"] += 1
            elif fa < fb:
                lpa, lpb = 0, 3
                stats[m["teamB"]]["w"] += 1
                stats[m["teamA"]]["l"] += 1
            else:
                lpa, lpb = 1, 1
                stats[m["teamA"]]["d"] += 1
                stats[m["teamB"]]["d"] += 1
            stats[m["teamA"]]["pts"] += lpa
            stats[m["teamB"]]["pts"] += lpb
            stats[m["teamA"]]["score"] += fa
            stats[m["teamB"]]["score"] += fb
            match_log.append(
                {
                    "teamA": m["teamA"],
                    "teamB": m["teamB"],
                    "leaguePtsA": lpa,
                    "leaguePtsB": lpb,
                    "fa": fa,
                    "fb": fb,
                }
            )
    rows = list(stats.values())
    rows.sort(key=cmp_to_key(lambda a, b: compare(a, b, match_log)))
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    return rows, match_log


final_rows, _match_log = compute_standings(38)

all_matches = []
for block in sorted(matches_by_gw, key=lambda x: int(x["gw"])):
    gw = int(block["gw"])
    for i, m in enumerate(block.get("matches") or [], 1):
        fa, fb = int(m["pointsA"]), int(m["pointsB"])
        if fa > fb:
            lpa, lpb, wynik, zwyciezca = 3, 0, "Wygrana A", m["teamA"]
            wynik_skrot = "1-0"
        elif fa < fb:
            lpa, lpb, wynik, zwyciezca = 0, 3, "Wygrana B", m["teamB"]
            wynik_skrot = "0-1"
        else:
            lpa, lpb, wynik, zwyciezca = 1, 1, "Remis", "—"
            wynik_skrot = "X"
        pa = players.get(m["teamA"], {})
        pb = players.get(m["teamB"], {})
        all_matches.append(
            {
                "gw": gw,
                "nr_meczu": i,
                "zespol_a": m["teamA"],
                "manager_a": pa.get("manager", ""),
                "discord_a": pa.get("discord", ""),
                "fpl_id_a": pa.get("id", ""),
                "pkt_fpl_a": fa,
                "pkt_h2h_a": lpa,
                "zespol_b": m["teamB"],
                "manager_b": pb.get("manager", ""),
                "discord_b": pb.get("discord", ""),
                "fpl_id_b": pb.get("id", ""),
                "pkt_fpl_b": fb,
                "pkt_h2h_b": lpb,
                "roznica_fpl": fa - fb,
                "wynik": wynik,
                "wynik_skrot": wynik_skrot,
                "zwyciezca": zwyciezca,
                "przegrany": (
                    m["teamB"] if fa > fb else m["teamA"] if fa < fb else "—"
                ),
            }
        )

mismatches = []
for r in final_rows:
    p = players.get(r["team"])
    if p and (
        p["rank"] != r["rank"] or p["pts"] != r["pts"] or p["score"] != r["score"]
    ):
        mismatches.append(
            (r["team"], r["rank"], p["rank"], r["pts"], p["pts"], r["score"], p["score"])
        )

print(f"players={len(players)} matches={len(all_matches)} mismatches={len(mismatches)}")
if mismatches:
    print("WARN mismatches:", mismatches[:8])

# ---- styles ----
thin = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
header_fill = PatternFill("solid", fgColor="0F172A")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
title_font = Font(name="Calibri", bold=True, size=16, color="0F172A")
subtitle_font = Font(name="Calibri", size=11, color="475569")
gold = PatternFill("solid", fgColor="FEF3C7")
silver = PatternFill("solid", fgColor="F1F5F9")
bronze = PatternFill("solid", fgColor="FFEDD5")
win_fill = PatternFill("solid", fgColor="D1FAE5")
draw_fill = PatternFill("solid", fgColor="FEF9C3")
loss_fill = PatternFill("solid", fgColor="FEE2E2")
alt_fill = PatternFill("solid", fgColor="F8FAFC")
accent = PatternFill("solid", fgColor="ECFDF5")
a_win_fill = PatternFill("solid", fgColor="E0F2FE")
b_win_fill = PatternFill("solid", fgColor="FCE7F3")
gw_header_fill = PatternFill("solid", fgColor="059669")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin


def autosize(ws, min_w=8, max_w=42):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_w, min(max_w, length + 2))


wb = Workbook()

# ---- Info ----
ws = wb.active
ws.title = "Info"
ws["A1"] = "Igrzyska Kapci Kłapcia — Sezon 2025/26"
ws["A1"].font = title_font
ws["A2"] = "Eksport archiwum ligi H2H (Skarb Kibica / FPL Arena)"
ws["A2"].font = subtitle_font
ws.merge_cells("A1:D1")
ws.merge_cells("A2:D2")

champ = final_rows[0]
info_rows = [
    ("Liga", "Igrzyska Kapci Kłapcia"),
    ("Sezon", "2025/26"),
    ("Format", "H2H · 1 liga · 20 drużyn · 38 kolejek"),
    ("Punktacja H2H", "Wygrana = 3 · Remis = 1 · Porażka = 0"),
    (
        "Tie-break tabeli",
        "1) Pkt H2H  2) Score FPL  3) Mini-liga H2H  4) Mini-liga Score  5) Nazwa",
    ),
    ("Źródło wyników", "wyniki_meczy.json"),
    ("Źródło profili", "src/data/players.ts"),
    ("Liczba meczów", len(all_matches)),
    ("Liczba kolejek", 38),
    (
        "Mistrz",
        f"{champ['team']} ({players.get(champ['team'], {}).get('manager', '')}) — {champ['pts']} pkt H2H / Score {champ['score']}",
    ),
    ("Data eksportu", datetime.now().strftime("%Y-%m-%d %H:%M")),
    (
        "Uwaga o rankingach",
        "Kolumny Poz./W/R/P/Pkt H2H/Score w arkuszu „Tabela końcowa” są PRZELICZONE z wyniki_meczy.json (jak w aplikacji Arena). "
        "W players.ts kilka dolnych pozycji ma lekko inne liczby redakcyjne — profil (menedżer, Discord, FPL ID, Super Star itd.) pochodzi z players.ts.",
    ),
]
for i, (k, v) in enumerate(info_rows, start=4):
    ws.cell(row=i, column=1, value=k).font = Font(bold=True)
    ws.cell(row=i, column=2, value=v)
    ws.cell(row=i, column=1).fill = accent

ws["A16"] = "Arkusze w pliku:"
ws["A16"].font = Font(bold=True, size=12)
for i, t in enumerate(
    [
        "1. Info — opis ligi i legenda",
        "2. Tabela końcowa — ranking po GW38 + dane profilowe",
        "3. Wyniki H2H — wszystkie mecze GW1–GW38 (menedżer, Discord, FPL ID, punkty)",
        "4. Wyniki wg kolejki — te same mecze pogrupowane GW1…GW38",
        "5. Bilans bezpośredni — macierz W-R-P kto vs kto",
    ],
    start=17,
):
    ws.cell(row=i, column=1, value=t)
autosize(ws)

# ---- Tabela końcowa ----
ws2 = wb.create_sheet("Tabela końcowa")
ws2["A1"] = "Tabela końcowa po GW38 — Igrzyska Kapci Kłapcia 2025/26"
ws2["A1"].font = title_font
ws2.merge_cells("A1:Z1")
ws2["A2"] = "Sortowanie jak w aplikacji: pkt H2H → Score FPL → bezpośrednie starcia"
ws2["A2"].font = subtitle_font

headers2 = [
    "Poz.",
    "Zespół",
    "Menedżer",
    "Discord",
    "FPL ID",
    "M",
    "W",
    "R",
    "P",
    "Pkt H2H",
    "Score FPL",
    "Poz. players.ts",
    "Śr. poz.",
    "Poz. po GW19",
    "Sezony FPL",
    "Best OR",
    "Sezon Best OR",
    "Transfery",
    "Hity (koszt)",
    "Zielone strzałki",
    "Pkt kapitana",
    "Najczęściej kapitan",
    "Najwięcej pkt (zawodnik)",
    "Super Star",
    "Rank Killer",
    "Pkt na ławce",
    "Seria wygranych",
    "Menedżer miesiąca",
    "Najlepsza GW",
    "Tygodnie #1",
    "Tygodnie ostatni",
]
hr = 4
for c, h in enumerate(headers2, 1):
    ws2.cell(row=hr, column=c, value=h)
style_header(ws2, hr, len(headers2))

numish = {1, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 18, 19, 20, 21, 26, 30, 31}
for i, r in enumerate(final_rows):
    p = players.get(r["team"], {})
    m_played = r["w"] + r["d"] + r["l"]
    row = [
        r["rank"],
        r["team"],
        p.get("manager", ""),
        p.get("discord", ""),
        p.get("id", ""),
        m_played,
        r["w"],
        r["d"],
        r["l"],
        r["pts"],
        r["score"],
        p.get("rank", ""),
        p.get("avgPosition", ""),
        p.get("gw19Rank", ""),
        p.get("seasons", ""),
        p.get("bestOr", ""),
        p.get("bestOrSeason", ""),
        p.get("transfers", ""),
        p.get("hits", ""),
        p.get("greenArrows", ""),
        p.get("captainPts", ""),
        p.get("mostCaptained", ""),
        p.get("mostPointsPlayer", ""),
        p.get("superStar", ""),
        p.get("rankKiller", ""),
        p.get("pointsBenched", ""),
        p.get("winStreak", ""),
        p.get("monthlyWins", ""),
        p.get("bestGw", ""),
        p.get("weeksTop", ""),
        p.get("weeksBottom", ""),
    ]
    rr = hr + 1 + i
    for c, val in enumerate(row, 1):
        cell = ws2.cell(row=rr, column=c, value=val)
        cell.border = thin
        cell.alignment = center if c in numish else left
        if i % 2 == 1:
            cell.fill = alt_fill
    fill = gold if r["rank"] == 1 else silver if r["rank"] == 2 else bronze if r["rank"] == 3 else None
    if fill:
        for c in range(1, 12):
            ws2.cell(row=rr, column=c).fill = fill

ws2.freeze_panes = "C5"
ws2.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers2))}{hr + len(final_rows)}"
autosize(ws2)

# ---- Wyniki H2H ----
ws3 = wb.create_sheet("Wyniki H2H")
ws3["A1"] = "Wszystkie wyniki H2H — GW1 do GW38"
ws3["A1"].font = title_font
ws3.merge_cells("A1:R1")
ws3["A2"] = f"Łącznie {len(all_matches)} meczów · FPL = pointsA/B · H2H = 3/1/0"
ws3["A2"].font = subtitle_font

headers3 = [
    "GW",
    "Nr meczu w GW",
    "Zespół A",
    "Menedżer A",
    "Discord A",
    "FPL ID A",
    "Pkt FPL A",
    "Pkt H2H A",
    "Zespół B",
    "Menedżer B",
    "Discord B",
    "FPL ID B",
    "Pkt FPL B",
    "Pkt H2H B",
    "Różnica FPL (A−B)",
    "Wynik",
    "Skrót",
    "Zwycięzca",
    "Przegrany",
]
hr = 4
for c, h in enumerate(headers3, 1):
    ws3.cell(row=hr, column=c, value=h)
style_header(ws3, hr, len(headers3))

for i, m in enumerate(all_matches):
    rr = hr + 1 + i
    row = [
        m["gw"],
        m["nr_meczu"],
        m["zespol_a"],
        m["manager_a"],
        m["discord_a"],
        m["fpl_id_a"],
        m["pkt_fpl_a"],
        m["pkt_h2h_a"],
        m["zespol_b"],
        m["manager_b"],
        m["discord_b"],
        m["fpl_id_b"],
        m["pkt_fpl_b"],
        m["pkt_h2h_b"],
        m["roznica_fpl"],
        m["wynik"],
        m["wynik_skrot"],
        m["zwyciezca"],
        m["przegrany"],
    ]
    if m["wynik"] == "Remis":
        fill = draw_fill
    elif m["wynik"] == "Wygrana A":
        fill = a_win_fill
    else:
        fill = b_win_fill
    for c, val in enumerate(row, 1):
        cell = ws3.cell(row=rr, column=c, value=val)
        cell.border = thin
        cell.alignment = center
        cell.fill = fill
    if m["wynik"] == "Wygrana A":
        ws3.cell(row=rr, column=3).fill = win_fill
        ws3.cell(row=rr, column=3).font = Font(bold=True)
    elif m["wynik"] == "Wygrana B":
        ws3.cell(row=rr, column=9).fill = win_fill
        ws3.cell(row=rr, column=9).font = Font(bold=True)

ws3.freeze_panes = "C5"
ws3.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers3))}{hr + len(all_matches)}"
autosize(ws3)

# ---- Wyniki wg kolejki ----
ws4 = wb.create_sheet("Wyniki wg kolejki")
ws4["A1"] = "Wyniki pogrupowane według kolejki"
ws4["A1"].font = title_font
row_i = 3
for gw in range(1, 39):
    gw_matches = [m for m in all_matches if m["gw"] == gw]
    if not gw_matches:
        continue
    cell = ws4.cell(row=row_i, column=1, value=f"GW{gw} — {len(gw_matches)} meczów")
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    for c in range(1, 8):
        ws4.cell(row=row_i, column=c).fill = gw_header_fill
    ws4.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=7)
    row_i += 1
    for c, h in enumerate(
        ["Zespół A", "Pkt FPL A", "Pkt H2H A", "Zespół B", "Pkt FPL B", "Pkt H2H B", "Zwycięzca"],
        1,
    ):
        cell = ws4.cell(row=row_i, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = center
    row_i += 1
    for m in gw_matches:
        vals = [
            m["zespol_a"],
            m["pkt_fpl_a"],
            m["pkt_h2h_a"],
            m["zespol_b"],
            m["pkt_fpl_b"],
            m["pkt_h2h_b"],
            m["zwyciezca"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws4.cell(row=row_i, column=c, value=v)
            cell.border = thin
            cell.alignment = center
            if m["wynik"] == "Remis":
                cell.fill = draw_fill
            elif m["wynik"] == "Wygrana A" and c == 1:
                cell.fill = win_fill
                cell.font = Font(bold=True)
            elif m["wynik"] == "Wygrana B" and c == 4:
                cell.fill = win_fill
                cell.font = Font(bold=True)
        row_i += 1
    row_i += 1
autosize(ws4)

# ---- Bilans bezpośredni ----
ws5 = wb.create_sheet("Bilans bezpośredni")
ws5["A1"] = "Bilans bezpośredni (W-R-P) — kto z kim w sezonie"
ws5["A1"].font = title_font
ws5["A2"] = "Komórka = bilans wiersza vs kolumna. Diagonal = — · zielony = przewaga wiersza"
ws5["A2"].font = subtitle_font

teams_ordered = [r["team"] for r in final_rows]
h2h = {a: {b: {"w": 0, "d": 0, "l": 0} for b in teams_ordered} for a in teams_ordered}
for m in all_matches:
    a, b = m["zespol_a"], m["zespol_b"]
    if a not in h2h or b not in h2h:
        continue
    if m["wynik"] == "Wygrana A":
        h2h[a][b]["w"] += 1
        h2h[b][a]["l"] += 1
    elif m["wynik"] == "Wygrana B":
        h2h[a][b]["l"] += 1
        h2h[b][a]["w"] += 1
    else:
        h2h[a][b]["d"] += 1
        h2h[b][a]["d"] += 1

ws5.cell(row=4, column=1, value="↓ vs →").font = Font(bold=True, color="FFFFFF")
ws5.cell(row=4, column=1).fill = header_fill
for j, t in enumerate(teams_ordered, 2):
    cell = ws5.cell(row=4, column=j, value=t)
    cell.font = Font(bold=True, color="FFFFFF", size=8)
    cell.fill = header_fill
    cell.alignment = Alignment(
        horizontal="center", vertical="center", textRotation=90, wrap_text=True
    )
    ws5.column_dimensions[get_column_letter(j)].width = 5

for i, a in enumerate(teams_ordered, 5):
    cell = ws5.cell(row=i, column=1, value=a)
    cell.font = Font(bold=True, color="FFFFFF", size=9)
    cell.fill = header_fill
    for j, b in enumerate(teams_ordered, 2):
        if a == b:
            val = "—"
            fill = PatternFill("solid", fgColor="E2E8F0")
        else:
            s = h2h[a][b]
            val = f"{s['w']}-{s['d']}-{s['l']}"
            if s["w"] > s["l"]:
                fill = win_fill
            elif s["w"] < s["l"]:
                fill = loss_fill
            elif s["w"] + s["d"] + s["l"] == 0:
                fill = alt_fill
                val = ""
            else:
                fill = draw_fill
        cell = ws5.cell(row=i, column=j, value=val)
        cell.fill = fill
        cell.alignment = center
        cell.border = thin
        cell.font = Font(size=8)

ws5.column_dimensions["A"].width = 22
ws5.row_dimensions[4].height = 80
ws5.freeze_panes = "B5"

wb.save(OUT)
print("SAVED:", OUT)
print("size_kb:", round(OUT.stat().st_size / 1024, 1))
