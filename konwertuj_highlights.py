"""
Konwertuje pliki highlights/{fpl_id}.xlsx (+ wyniki H2H) do player_highlights.json.
Jeden plik xlsx = jeden menedżer (ID w arkuszu Team Info lub nazwie pliku).
"""
import json
import os
import glob
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_JSON = os.path.join(BASE_DIR, "player_highlights.json")
WYNIKI_JSON = os.path.join(BASE_DIR, "wyniki_meczy.json")
HIGHLIGHTS_DIR = os.path.join(BASE_DIR, "highlights")

PLAYERS = [
    {"id": 22952, "team": "Kapcie Kłapcia"},
    {"id": 49321, "team": "świnie Pepa"},
    {"id": 43986, "team": "Bad Kompany"},
    {"id": 298030, "team": "Pewniaczki"},
    {"id": 425158, "team": "Scuderia Blaugrana"},
    {"id": 418929, "team": "Pusan"},
    {"id": 3804416, "team": "VeB"},
    {"id": 55981, "team": "FC Defconiarze"},
    {"id": 4002, "team": "Furiosa FC Morfeusza"},
    {"id": 1109898, "team": "immigrants fc"},
    {"id": 126745, "team": "Przemsza Klucze"},
    {"id": 248187, "team": "C30-C39"},
    {"id": 2953280, "team": "Bolesławiec King"},
    {"id": 9084, "team": "Wirtz Team Ever"},
    {"id": 1178442, "team": "Boom Saka Laka"},
    {"id": 546068, "team": "MQUC"},
    {"id": 68435, "team": "FcpoNalewce"},
    {"id": 24962, "team": "Ulane Warchlaki"},
    {"id": 3873739, "team": "Jarząbki"},
    {"id": 3749264, "team": "MnstrNaf"},
]

CHIP_LABELS = {
    "bboost": "Bench Boost (punkty z ławki)",
    "wildcard": "Wildcard (pełna przebudowa)",
    "3xc": "Potrójny Kapitan",
    "freehit": "Free Hit (drużyna za 0 kosztów)",
}

POSITION_LABELS = {
    1: "Bramkarze",
    2: "Obrońcy",
    3: "Pomocnicy",
    4: "Napastnicy",
}


def aggregate_squad_from_gw_picks(wb):
    """Suma punktów z GW Picks dla każdego zawodnika (skład + ławka)."""
    if "GW Picks" not in wb.sheetnames:
        return {}
    by_id = {}
    for r in sheet_rows(wb, "GW Picks"):
        pid = str(r.get("element", "")).strip()
        if not pid.isdigit():
            continue
        pts = float(r.get("points") or 0)
        eff = float(r.get("eff_points") or 0)
        mult = int(r.get("multiplier") or 1)
        et = int(r.get("element_type") or 0)
        name = r.get("name")
        rec = by_id.setdefault(pid, {
            "name": name or id_to_name_placeholder(pid),
            "elementId": int(pid),
            "position": et,
            "pointsBase": 0.0,
            "captainBonus": 0.0,
            "tcBonus": 0.0,
            "points": 0.0,
        })
        if name:
            rec["name"] = name
        if et:
            rec["position"] = et
        rec["pointsBase"] += pts
        rec["points"] += eff
        if mult == 2:
            rec["captainBonus"] += pts
        elif mult == 3:
            rec["tcBonus"] += pts * 2
    for rec in by_id.values():
        rec["pointsBase"] = int(round(rec["pointsBase"]))
        rec["captainBonus"] = int(round(rec["captainBonus"]))
        rec["tcBonus"] = int(round(rec["tcBonus"]))
        rec["points"] = int(round(rec["points"]))
    return by_id


def id_to_name_placeholder(pid):
    return f"ID {pid}"


def sheet_rows(wb, name):
    ws = wb[name]
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def team_info_dict(wb):
    info = {}
    for row in wb["Team Info"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            info[row[0]] = row[1]
    return info


def extract_full(wb):
    info = team_info_dict(wb)
    fpl_id = int(info.get("id", 0))

    gw_rows = sheet_rows(wb, "GW Info")
    gw_points = []
    chips = []
    for r in gw_rows:
        gw = int(r["gw"])
        pts = int(r["points"] or 0)
        gw_points.append({
            "gw": gw,
            "points": pts,
            "bench": int(r.get("points_on_bench") or 0),
            "transfers": int(r.get("event_transfers") or 0),
            "hitCost": int(r.get("event_transfers_cost") or 0),
            "overallRank": int(r["overall_rank"]) if r.get("overall_rank") else None,
        })
        if r.get("active_chip"):
            chips.append({
                "gw": gw,
                "chip": r["active_chip"],
                "chipLabel": CHIP_LABELS.get(r["active_chip"], str(r["active_chip"])),
                "points": pts,
            })

    gw_points.sort(key=lambda x: x["gw"])
    sorted_pts = sorted(gw_points, key=lambda x: x["points"], reverse=True)
    best_gws = [{"gw": g["gw"], "points": g["points"]} for g in sorted_pts[:3]]
    worst_gws = [{"gw": g["gw"], "points": g["points"]} for g in sorted_pts[-3:]]

    gain_loss = sheet_rows(wb, "Player Gain - Loss")
    gain_loss.sort(key=lambda x: float(x.get("net") or 0), reverse=True)
    top_gains = [
        {"name": r["name"], "net": round(float(r["net"]), 1), "points": int(r.get("points") or 0)}
        for r in gain_loss[:5]
    ]
    top_losses = [
        {"name": r["name"], "net": round(float(r["net"]), 1), "points": int(r.get("points") or 0)}
        for r in gain_loss[-5:][::-1]
    ]

    transfers = [r for r in sheet_rows(wb, "Transfer Quality") if r.get("sold_name") and r.get("bought_name")]
    transfers_valid = [t for t in transfers if t.get("rp_diff") is not None]
    best_transfers = sorted(transfers_valid, key=lambda x: float(x["rp_diff"]), reverse=True)[:4]
    worst_transfers = sorted(transfers_valid, key=lambda x: float(x["rp_diff"]))[:4]

    def fmt_transfer(t):
        return {
            "gw": int(t["gw"]),
            "sold": t.get("sold_name"),
            "bought": t.get("bought_name"),
            "rpDiff": round(float(t["rp_diff"]), 1),
            "xpDiff": round(float(t["xp_diff"]), 1) if t.get("xp_diff") is not None else None,
        }

    gain_moments = []
    for r in sheet_rows(wb, "Gain Analysis"):
        if not r.get("name") or r.get("pts") is None or r.get("xpts") is None:
            continue
        diff = float(r["pts"]) - float(r["xpts"])
        gain_moments.append({
            "name": r["name"], "gw": int(r["gw"]), "points": int(r["pts"]),
            "diff": round(diff, 1),
        })
    gain_moments.sort(key=lambda x: x["diff"], reverse=True)

    loss_moments = []
    for r in sheet_rows(wb, "Loss Analysis"):
        if not r.get("name") or r.get("pts") is None or r.get("xpts") is None:
            continue
        diff = float(r["pts"]) - float(r["xpts"])
        loss_moments.append({
            "name": r["name"], "gw": int(r["gw"]), "points": int(r["pts"]),
            "diff": round(diff, 1),
        })
    loss_moments.sort(key=lambda x: x["diff"])

    formations = []
    for r in sheet_rows(wb, "Formations"):
        if r.get("formation"):
            formations.append({
                "name": r["formation"],
                "count": int(r.get("count") or 0),
                "avg": round(float(r.get("avg") or 0), 1),
            })
    formations.sort(key=lambda x: x["count"], reverse=True)

    point_sources = []
    for r in sheet_rows(wb, "Event Points Total"):
        t = r.get("type")
        if not t:
            continue
        pts = int(r.get("points") or 0)
        pct_raw = r.get("percentage")
        pct = None
        if pct_raw not in (None, "-", ""):
            try:
                pct = float(str(pct_raw).replace("%", ""))
            except ValueError:
                pct = None
        cnt = r.get("count")
        point_sources.append({
            "type": t,
            "points": pts,
            "pct": pct,
            "count": cnt if cnt not in (None, "-") else None,
        })
    point_sources.sort(key=lambda x: abs(x["points"]), reverse=True)

    points_by_position = []
    by_element = {1: 0.0, 2: 0.0, 3: 0.0, 4: 0.0}
    if "GW Picks" in wb.sheetnames:
        for row in wb["GW Picks"].iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            et = row[5]
            if et not in by_element:
                continue
            raw_pts = row[13] if row[13] is not None else row[12]
            by_element[et] += float(raw_pts or 0)
    pos_total = sum(by_element.values())
    for et in (1, 2, 3, 4):
        pts = by_element[et]
        points_by_position.append({
            "position": et,
            "label": POSITION_LABELS[et],
            "points": int(round(pts)),
            "pct": round(pts / pos_total * 100, 1) if pos_total else 0,
        })
    points_by_position.sort(key=lambda x: x["points"], reverse=True)

    exp_rows = sheet_rows(wb, "Exp vs Real Performance")
    over = sum(1 for r in exp_rows if float(r.get("diff") or 0) > 0)
    under = sum(1 for r in exp_rows if float(r.get("diff") or 0) < 0)
    exp_sorted = sorted(exp_rows, key=lambda x: float(x.get("diff") or 0), reverse=True)
    best_surprise = None
    worst_surprise = None
    if exp_sorted:
        b = exp_sorted[0]
        w = exp_sorted[-1]
        best_surprise = {"gw": int(b["gw"]), "diff": round(float(b["diff"]), 1), "points": int(b["rp"])}
        worst_surprise = {"gw": int(w["gw"]), "diff": round(float(w["diff"]), 1), "points": int(w["rp"])}

    targets = []
    for r in sheet_rows(wb, "Target Event Ratios"):
        if r.get("target"):
            targets.append({
                "label": r["target"],
                "ratio": round(float(r.get("success_ratio") or 0), 1),
                "info": r.get("info"),
            })

    id_to_name = {str(r.get("id")): r["name"] for r in gain_loss if r.get("id")}
    squad_by_id = aggregate_squad_from_gw_picks(wb)

    dream_team = []
    for r in sheet_rows(wb, "Team of the Year"):
        if not r.get("lineup"):
            continue
        pid = str(r.get("id", "")).strip()
        pos = int(r.get("position") or 0)
        squad_stats = squad_by_id.get(pid, {})
        dream_team.append({
            "name": r.get("name") or id_to_name.get(pid, f"ID {pid}"),
            "elementId": int(pid) if pid.isdigit() else None,
            "position": pos,
            "posLabel": r.get("pos"),
            "points": int(r.get("points_total") or squad_stats.get("points") or 0),
            "pointsBase": squad_stats.get("pointsBase", 0),
            "captainBonus": squad_stats.get("captainBonus", 0),
            "tcBonus": squad_stats.get("tcBonus", 0),
            "captaincies": int(r.get("cap_count") or 0),
            "pitchX": float(r.get("x") or 0),
            "pitchY": float(r.get("y") or 0),
        })
    dream_team.sort(key=lambda x: (-x["pitchY"], x["pitchX"]))

    squad_players = sorted(squad_by_id.values(), key=lambda x: (-x["points"], x["name"]))

    total_bench = sum(g["bench"] for g in gw_points)
    total_hits = sum(g["hitCost"] for g in gw_points)
    avg_points = round(sum(g["points"] for g in gw_points) / max(len(gw_points), 1), 1)

    season_or = info.get("summary_overall_rank")
    if season_or is not None:
        season_or = int(season_or)

    return {
        "source": "full",
        "fplId": fpl_id,
        "seasonOr": season_or,
        "avgGwPoints": avg_points,
        "totalBench": total_bench,
        "totalHitCost": total_hits,
        "gwPoints": gw_points,
        "chips": chips,
        "bestGWs": best_gws,
        "worstGWs": worst_gws,
        "topGains": top_gains,
        "topLosses": top_losses,
        "bestTransfers": [fmt_transfer(t) for t in best_transfers],
        "worstTransfers": [fmt_transfer(t) for t in worst_transfers],
        "gainMoments": gain_moments[:5],
        "lossMoments": loss_moments[:5],
        "formations": formations[:5],
        "pointSources": point_sources,
        "pointsByPosition": points_by_position,
        "expSummary": {"overperform": over, "underperform": under, "best": best_surprise, "worst": worst_surprise},
        "targetRatios": targets,
        "dreamTeam": dream_team[:11],
        "squadPlayers": squad_players,
    }


def load_h2h_by_team():
    if not os.path.exists(WYNIKI_JSON):
        return {}
    with open(WYNIKI_JSON, encoding="utf-8") as f:
        data = json.load(f)
    by_team = {p["team"]: p["id"] for p in PLAYERS}
    result = {pid: [] for pid in by_team.values()}
    for block in data:
        gw = block["gw"]
        for m in block.get("matches", []):
            for side, team_key, pts_key, opp_key, opp_pts_key in [
                ("home", "teamA", "pointsA", "teamB", "pointsB"),
                ("away", "teamB", "pointsB", "teamA", "pointsA"),
            ]:
                team = m[team_key]
                if team not in by_team:
                    continue
                pid = by_team[team]
                gf, ga = int(m[pts_key]), int(m[opp_pts_key])
                outcome = "W" if gf > ga else ("L" if gf < ga else "D")
                result[pid].append({
                    "gw": gw,
                    "fplPoints": gf,
                    "opponent": m[opp_key],
                    "oppPoints": ga,
                    "outcome": outcome,
                    "margin": gf - ga,
                    "isHome": side == "home",
                })
    for pid in result:
        result[pid].sort(key=lambda x: x["gw"])
    return result


def streaks(outcomes):
    best_w = best_l = cur_w = cur_l = 0
    for o in outcomes:
        if o == "W":
            cur_w += 1
            cur_l = 0
        elif o == "L":
            cur_l += 1
            cur_w = 0
        else:
            cur_w = cur_l = 0
        best_w = max(best_w, cur_w)
        best_l = max(best_l, cur_l)
    return best_w, best_l


def build_basic(fpl_id, team, h2h_rows):
    if not h2h_rows:
        return {"source": "basic", "fplId": fpl_id, "gwPoints": [], "h2h": []}

    gw_points = [{"gw": r["gw"], "points": r["fplPoints"]} for r in h2h_rows]
    sorted_pts = sorted(gw_points, key=lambda x: x["points"], reverse=True)
    outcomes = [r["outcome"] for r in h2h_rows]
    win_s, loss_s = streaks(outcomes)
    last5 = outcomes[-5:]
    first_half = gw_points[:19]
    second_half = gw_points[19:]
    avg1 = round(sum(g["points"] for g in first_half) / max(len(first_half), 1), 1)
    avg2 = round(sum(g["points"] for g in second_half) / max(len(second_half), 1), 1)
    margins = sorted(h2h_rows, key=lambda x: x["margin"], reverse=True)

    return {
        "source": "basic",
        "fplId": fpl_id,
        "avgGwPoints": round(sum(g["points"] for g in gw_points) / len(gw_points), 1),
        "gwPoints": gw_points,
        "h2h": h2h_rows,
        "bestGWs": [{"gw": g["gw"], "points": g["points"]} for g in sorted_pts[:3]],
        "worstGWs": [{"gw": g["gw"], "points": g["points"]} for g in sorted_pts[-3:]],
        "h2hStreaks": {"maxWins": win_s, "maxLosses": loss_s, "last5": last5},
        "seasonSplit": {"firstHalfAvg": avg1, "secondHalfAvg": avg2, "trend": avg2 - avg1},
        "biggestWin": {
            "gw": margins[0]["gw"],
            "margin": margins[0]["margin"],
            "opponent": margins[0]["opponent"],
            "score": f"{margins[0]['fplPoints']}:{margins[0]['oppPoints']}",
        } if margins and margins[0]["margin"] > 0 else None,
        "heaviestLoss": {
            "gw": margins[-1]["gw"],
            "margin": margins[-1]["margin"],
            "opponent": margins[-1]["opponent"],
            "score": f"{margins[-1]['fplPoints']}:{margins[-1]['oppPoints']}",
        } if margins and margins[-1]["margin"] < 0 else None,
    }


def merge_h2h_into_full(full, h2h_rows):
    h2h_by_gw = {r["gw"]: r for r in h2h_rows}
    for g in full.get("gwPoints", []):
        h = h2h_by_gw.get(g["gw"])
        if h:
            g["h2hOutcome"] = h["outcome"]
            g["opponent"] = h["opponent"]
            g["h2hScore"] = f"{h['fplPoints']}:{h['oppPoints']}"

    outcomes = [r["outcome"] for r in h2h_rows]
    win_s, loss_s = streaks(outcomes)
    full["h2hStreaks"] = {"maxWins": win_s, "maxLosses": loss_s, "last5": outcomes[-5:]}
    full["h2h"] = h2h_rows

    if h2h_rows:
        margins = sorted(h2h_rows, key=lambda x: x["margin"], reverse=True)
        if margins[0]["margin"] > 0:
            b = margins[0]
            full["biggestWin"] = {
                "gw": b["gw"],
                "margin": b["margin"],
                "opponent": b["opponent"],
                "score": f"{b['fplPoints']}:{b['oppPoints']}",
            }
        if margins[-1]["margin"] < 0:
            l = margins[-1]
            full["heaviestLoss"] = {
                "gw": l["gw"],
                "margin": l["margin"],
                "opponent": l["opponent"],
                "score": f"{l['fplPoints']}:{l['oppPoints']}",
            }
        pts = full.get("gwPoints", [])
        if len(pts) >= 19:
            first = pts[:19]
            second = pts[19:]
            avg1 = round(sum(g["points"] for g in first) / len(first), 1)
            avg2 = round(sum(g["points"] for g in second) / len(second), 1)
            full["seasonSplit"] = {"firstHalfAvg": avg1, "secondHalfAvg": avg2, "trend": round(avg2 - avg1, 1)}

    return full


def find_xlsx_files():
    if not os.path.isdir(HIGHLIGHTS_DIR):
        return []
    return sorted(glob.glob(os.path.join(HIGHLIGHTS_DIR, "*.xlsx")))


def main():
    h2h_all = load_h2h_by_team()
    full_by_id = {}

    for path in find_xlsx_files():
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            full = extract_full(wb)
            fid = full["fplId"]
            if fid:
                full_by_id[fid] = full
                print(f"  full: {path} -> ID {fid}")
        except Exception as e:
            print(f"  blad {path}: {e}")

    output = {}
    for p in PLAYERS:
        pid = p["id"]
        h2h = h2h_all.get(pid, [])
        if pid in full_by_id:
            output[str(pid)] = merge_h2h_into_full(full_by_id[pid], h2h)
        else:
            output[str(pid)] = build_basic(pid, p["team"], h2h)

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    full_count = sum(1 for v in output.values() if v.get("source") == "full")
    print(f"Zapisano {OUTPUT_JSON}: {len(output)} graczy ({full_count} pelnych analiz Excel)")


if __name__ == "__main__":
    print("Generowanie player_highlights.json...")
    main()
