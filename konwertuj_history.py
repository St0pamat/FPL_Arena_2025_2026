"""
Konwertuje highlights/*_history25-26.xlsx do player_season_history.json.
Dane FPL globalne (vs Top 10k, ranki GW, rotacja składu) — uzupełnienie profili.
"""
import glob
import json
import os
import re

import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
HIGHLIGHTS_DIR = os.path.join(BASE_DIR, "highlights")
OUT_ROOT = os.path.join(BASE_DIR, "player_season_history.json")
OUT_PUBLIC = os.path.join(BASE_DIR, "public", "player_season_history.json")


def find_gw_sheet(wb, pid):
    target = f"{pid} GW History"
    if target in wb.sheetnames:
        return target
    for s in wb.sheetnames:
        if "GW History" in s:
            return s
    return None


def parse_name_count(raw):
    if not raw:
        return None, None
    s = str(raw).strip()
    m = re.match(r"^(.+?)\s+(\d+)$", s)
    if m:
        return m.group(1).strip(), int(m.group(2))
    return s, None


def extract(path):
    fname = os.path.basename(path)
    pid = int(re.match(r"(\d+)", fname).group(1))
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = find_gw_sheet(wb, pid)
    if not sheet:
        wb.close()
        return None

    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    summary = {}
    if rows and rows[0]:
        for i, k in enumerate(rows[0]):
            if k and i < len(rows[1] or ()):
                summary[k] = rows[1][i]

    gw_hdr = None
    for i, row in enumerate(rows):
        if row and row[0] == "GW":
            gw_hdr = i
            break

    gw_rows = []
    if gw_hdr is not None:
        headers = rows[gw_hdr]
        for row in rows[gw_hdr + 1 :]:
            if not row or row[0] is None:
                continue
            try:
                int(row[0])
            except (TypeError, ValueError):
                continue
            rec = dict(zip(headers, row))
            gw_rows.append({
                "gw": int(rec["GW"]),
                "points": int(rec.get("Points") or 0),
                "top10kAvg": round(float(rec.get("Top10k Avg") or 0), 2),
                "overallAvg": round(float(rec.get("Overall Avg") or 0), 2),
                "vsTop10k": round(float(rec.get("Points") or 0) - float(rec.get("Top10k Avg") or 0), 2),
                "bench": int(rec.get("Bench") or 0),
                "gwRank": int(rec["GW Rank"]) if rec.get("GW Rank") else None,
                "captain": str(rec["Captain"]).strip() if rec.get("Captain") else None,
                "mainGain": str(rec["Main Gain"]).strip() if rec.get("Main Gain") else None,
                "mainDamage": str(rec["Main Damage"]).strip() if rec.get("Main Damage") else None,
                "teamValue": float(rec["Team Value"]) if rec.get("Team Value") is not None else None,
                "chip": str(rec["Chip"]).strip() if rec.get("Chip") and rec.get("Chip") != "None" else None,
            })

    gw_rows.sort(key=lambda x: x["gw"])
    gw_rows = [g for g in gw_rows if 1 <= g["gw"] <= 38]

    vs_top10k = [g["vsTop10k"] for g in gw_rows]
    most_started_name, most_started_count = parse_name_count(summary.get("Most started"))
    most_benched_name, most_benched_count = parse_name_count(summary.get("Most Benched"))

    best_gw_rank = min((g["gwRank"] for g in gw_rows if g["gwRank"]), default=None)
    worst_gw_rank = max((g["gwRank"] for g in gw_rows if g["gwRank"]), default=None)

    peak_value = max((g["teamValue"] for g in gw_rows if g["teamValue"]), default=None)

    return str(pid), {
        "fplId": pid,
        "seasonOr": int(summary["OR"]) if summary.get("OR") else None,
        "fplTotalPoints": int(summary["Points"]) if summary.get("Points") else sum(g["points"] for g in gw_rows),
        "avgVsTop10k": round(sum(vs_top10k) / len(vs_top10k), 2) if vs_top10k else None,
        "avgVsOverall": round(
            sum(g["points"] - g["overallAvg"] for g in gw_rows) / len(gw_rows), 2
        ) if gw_rows else None,
        "weeksAboveTop10k": sum(1 for v in vs_top10k if v > 0),
        "weeksBelowTop10k": sum(1 for v in vs_top10k if v < 0),
        "weeksEqualTop10k": sum(1 for v in vs_top10k if v == 0),
        "bestGwRank": best_gw_rank,
        "worstGwRank": worst_gw_rank,
        "peakTeamValue": round(peak_value, 1) if peak_value else None,
        "mostStarted": {"name": most_started_name, "count": most_started_count},
        "mostBenchedPlayer": {"name": most_benched_name, "count": most_benched_count},
        "gwDetails": gw_rows,
    }


def main():
    output = {}
    files = sorted(glob.glob(os.path.join(HIGHLIGHTS_DIR, "*history*.xlsx")))
    for path in files:
        try:
            parsed = extract(path)
            if parsed:
                pid, data = parsed
                output[pid] = data
                print(f"  OK {os.path.basename(path)}")
        except Exception as e:
            print(f"  ERR {path}: {e}")

    for out_path in (OUT_ROOT, OUT_PUBLIC):
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"Zapisano {len(output)} profili -> player_season_history.json")


if __name__ == "__main__":
    print("Generowanie player_season_history.json...")
    main()
