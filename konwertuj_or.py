"""
Łączy historyczny OR z Akta Gladiatorów (przed sezonem 2025/26)
z OR końcowym sezonu 2025/26 z plików highlights/*.xlsx → gladiator_or.json
"""
import json
import os
import glob
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_JSON = os.path.join(BASE_DIR, "gladiator_or.json")
HIGHLIGHTS_DIR = os.path.join(BASE_DIR, "highlights")


def find_akta_path():
    for name in os.listdir(BASE_DIR):
        if name.endswith(".xlsx") and "Akta" in name:
            return os.path.join(BASE_DIR, name)
    return None


def load_historical():
    path = find_akta_path()
    if not path:
        return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    out = {}
    for r in range(2, ws.max_row + 1):
        fid = ws.cell(r, 9).value
        if not fid:
            continue
        fid = int(fid)
        orv = ws.cell(r, 13).value
        season = ws.cell(r, 14).value
        hist_or = int(orv) if orv not in (None, "", 0) else None
        out[str(fid)] = {
            "historicalOr": hist_or,
            "historicalOrSeason": str(season).strip() if season else None,
        }
    return out


def load_season_or():
    out = {}
    if not os.path.isdir(HIGHLIGHTS_DIR):
        return out
    for path in glob.glob(os.path.join(HIGHLIGHTS_DIR, "*.xlsx")):
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            info = {
                row[0]: row[1]
                for row in wb["Team Info"].iter_rows(min_row=2, values_only=True)
                if row[0]
            }
            fid = int(info.get("id", 0) or 0)
            rank = info.get("summary_overall_rank")
            if fid and rank is not None:
                out[str(fid)] = int(rank)
        except Exception as e:
            print(f"  blad {path}: {e}")
    return out


def main():
    historical = load_historical()
    season = load_season_or()
    merged = {}
    all_ids = set(historical.keys()) | set(season.keys())
    for fid in all_ids:
        merged[fid] = {
            **historical.get(fid, {"historicalOr": None, "historicalOrSeason": None}),
            "seasonOr": season.get(fid),
        }
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)
    print(f"Zapisano {OUTPUT_JSON}: {len(merged)} graczy")


if __name__ == "__main__":
    print("Generowanie gladiator_or.json...")
    main()
